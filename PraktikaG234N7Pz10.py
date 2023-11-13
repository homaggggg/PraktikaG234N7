import pandas as pd
from openpyxl import load_workbook

file = 'Машина 1.xlsx'
wb = load_workbook(file)
sheet = wb.get_sheet_by_name('Лист1')

#for i in range(2, 5):
#     print(i, sheet.cell(row=i, column=2).value)

Crepesh = {'Болт', 'Гайка', 'Штифт', 'Шайба', 'Шуруп'}
Decor = {'Коврики', 'Подушки'}
engletters = 'QWERTYUIOPASDFGHJKLZXCVBNM'

#задание 3
for i in range(2, sheet.max_row+1):
#    print(i, sheet.cell(row=i, column=6).value, sheet.cell(row=i, column=8).value)
#    print(i, sheet.cell(row=i, column=5).value)
    #задание 3
    if sheet.cell(row=i, column=8).value == 'Сборка':
        sheet.cell(row=i, column=9).value = 'Сборка'
#            print('Сборка')
    elif sheet.cell(row=i, column=6).value in Crepesh:
#            print('Крепеж')
        sheet.cell(row=i, column=9).value = 'Крепеж'
    elif sheet.cell(row=i, column=6).value in Decor:
#            print('Декор')
        sheet.cell(row=i, column=9).value = 'Декор'
    else:
#            print('Детали')
        sheet.cell(row=i, column=9).value = 'Детали'

        #задание 4
    if sheet.cell(row=i, column=5).value.split('-')[0] == "ГОСТ":
#            print('ГОСТ РФ')
        sheet.cell(row=i, column=10).value = 'ГОСТ РФ'
    elif sheet.cell(row=i, column=5).value[0] in "ЙЦУКЕНГШЩЗХЪФЫВАПРОЛДЖЭЯЧСМИТЬБЮ":
#            print('РФ')
        sheet.cell(row=i, column=10).value = 'РФ'
    else:
#            print('Иностранное')
        sheet.cell(row=i, column=10).value = 'Иностранное'

    print(i, sheet.cell(row=i, column=6).value, sheet.cell(row=i, column=8).value, sheet.cell(row=i, column=9).value, sheet.cell(row=i, column=10).value)

wb.save('Машина 1-разметка.xlsx') 
        

#задание 4



#writer = pd.ExcelWriter('Машина 1.xlsx', engine='xlsxwriter')
#df1.to_excel(writer, 'Лист1')
#writer.save()