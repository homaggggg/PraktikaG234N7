import pandas as pd
from openpyxl import load_workbook

wb_1 = load_workbook('Машина 1.xlsx')
wb_2 = load_workbook('Машина 2.xlsx')
sheet_1 = wb_1.get_sheet_by_name('Лист1')
sheet_2 = wb_2.get_sheet_by_name('Лист1')

#for i in range(2, 5):
#     print(i, sheet.cell(row=i, column=2).value)

Crepesh = {'Болт', 'Гайка', 'Штифт', 'Шайба', 'Шуруп'}
Decor = {'Коврики', 'Подушки'}
engletters = 'QWERTYUIOPASDFGHJKLZXCVBNM'

#задание 3
for i in range(2, sheet_2.max_row+1):
#    print(i, sheet.cell(row=i, column=6).value, sheet.cell(row=i, column=8).value)
#    print(i, sheet.cell(row=i, column=5).value)
    #задание 3
    for j in range(2, sheet_2.max_row+1):
        if sheet_2.cell(row=i, column=5).value == sheet_1.cell(row=j, column=5).value:
            if sheet_2.cell(row=i, column=7).value != sheet_1.cell(row=j, column=7).value:
                sheet_2.cell(row=i, column=8).value = 'Изменилось количество'
            else:
                sheet_2.cell(row=i, column=8).value = 'Элемент не изменен'    
    
    if sheet_2.cell(row=i, column=8).value == None:
        sheet_2.cell(row=i, column=8).value = 'Элемент добавлен'
    print(i, sheet_2.cell(row=i, column=8).value)

wb_2.save('Машина 2-разметка.xlsx') 
        

#задание 4



#writer = pd.ExcelWriter('Машина 1.xlsx', engine='xlsxwriter')
#df1.to_excel(writer, 'Лист1')
#writer.save()